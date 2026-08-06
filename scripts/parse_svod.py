# -*- coding: utf-8 -*-
"""
Разбор листа «Общий свод 2026» в нормализованную структуру.

Логика листа-источника (заполняется агрономами ежедневно):
  строка 3            — шапка; со столбца N вправо идут даты сезона
  столбец A           — название операции (вида работ) ИЛИ подоперации
  столбец B           — культура
  столбец C           — предприятие
  столбец D           — общий план, га
  столбцы J..M        — плановые и фактические сроки начала/окончания
  столбцы N..         — выработка за день, га

Классификация строк:
  A заполнен, B пуст, D пуст            -> заголовок раздела (напр. «ВЕСЕННЕ-ПОЛЕВЫЕ РАБОТЫ»)
  A заполнен, B пуст, D = формула SUM   -> строка-итог вида работ (группа)
  A заполнен, B заполнен                -> строка данных с собственной подоперацией
  A пуст,     B заполнен                -> строка данных, наследует вид работ у группы
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

SHEET = "Общий свод 2026"
HEADER_ROW = 3
FIRST_DAY_COL = 14  # N
COL = {"op": 1, "culture": 2, "farm": 3, "plan": 4,
       "plan_start": 10, "plan_end": 11, "fact_start": 12, "fact_end": 13}


def clean(value) -> str:
    """Нормализует текст ячейки: убирает переносы, двойные пробелы, NBSP."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def as_number(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        text = value.replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def slugify(text: str) -> str:
    table = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
        "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    out = "".join(table.get(ch, ch) for ch in text.lower())
    out = re.sub(r"[^a-z0-9]+", "-", out).strip("-")
    return out or "op"


@dataclass
class Row:
    """Одна строка данных = срез «вид работ × культура × предприятие»."""
    section: str
    op: str
    sub: str
    culture: str
    farm: str
    plan: float
    plan_start: dt.date | None
    plan_end: dt.date | None
    fact_start: dt.date | None
    fact_end: dt.date | None
    daily: dict = field(default_factory=dict)
    excel_row: int = 0

    @property
    def fact(self) -> float:
        return round(sum(self.daily.values()), 2)


def read_dates(ws) -> dict:
    """Возвращает {номер столбца: дата} для календарной части листа."""
    dates = {}
    for col in range(FIRST_DAY_COL, ws.max_column + 1):
        day = as_date(ws.cell(HEADER_ROW, col).value)
        if day:
            dates[col] = day
    return dates


def parse(path: str | Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"В книге нет листа «{SHEET}». Найдено: {wb.sheetnames}")
    ws = wb[SHEET]

    dates = read_dates(ws)
    if not dates:
        raise SystemExit("Не найдены столбцы с датами в строке 3 (начиная со столбца N).")

    rows: list[Row] = []
    section = ""
    current_op = ""
    started = False  # данные начинаются после первого заголовка раздела (шапка с погодой — выше)

    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        op_cell = clean(ws.cell(r, COL["op"]).value)
        culture = clean(ws.cell(r, COL["culture"]).value)
        farm = clean(ws.cell(r, COL["farm"]).value)
        plan_raw = ws.cell(r, COL["plan"]).value
        plan = as_number(plan_raw)
        is_formula = isinstance(plan_raw, str) and plan_raw.startswith("=")

        daily = {}
        for col, day in dates.items():
            val = as_number(ws.cell(r, col).value)
            if val:
                daily[day.isoformat()] = round(val, 2)

        # Заголовок раздела: только текст, ни плана, ни выработки
        if op_cell and not culture and plan is None and not is_formula and not daily:
            if op_cell.upper() == op_cell and len(op_cell) > 3:
                section = op_cell
                started = True
                current_op = ""
            continue

        if not started:
            continue  # блок погоды и прочая шапка

        # Строка-итог вида работ: план собирается формулой SUM по дочерним строкам
        if op_cell and not culture and is_formula:
            current_op = op_cell
            continue

        # Пустая строка
        if not culture and not daily and plan is None:
            continue

        # Строка данных
        if op_cell and culture:
            # у строки собственное имя (подоперация внутри вида работ)
            op_name = current_op or op_cell
            sub = op_cell
        else:
            op_name = current_op or "Прочие работы"
            sub = ""

        rows.append(Row(
            section=section,
            op=op_name,
            sub=sub,
            culture=culture or "—",
            farm=farm or "—",
            plan=plan or 0.0,
            plan_start=as_date(ws.cell(r, COL["plan_start"]).value),
            plan_end=as_date(ws.cell(r, COL["plan_end"]).value),
            fact_start=as_date(ws.cell(r, COL["fact_start"]).value),
            fact_end=as_date(ws.cell(r, COL["fact_end"]).value),
            daily=daily,
            excel_row=r,
        ))

    return {"rows": rows, "dates": sorted({d.isoformat() for d in dates.values()})}


def status(pct: float) -> str:
    """Светофор. Пороги согласованы с примерами вкладок в исходной книге."""
    if pct >= 99.5:
        return "done"
    if pct >= 70:
        return "ok"
    if pct >= 35:
        return "warn"
    return "risk"


def build_model(parsed: dict, report_date: dt.date | None = None) -> dict:
    rows: list[Row] = parsed["rows"]

    active_days = sorted({d for row in rows for d in row.daily})
    if report_date is None:
        report_date = dt.date.fromisoformat(active_days[-1]) if active_days else dt.date.today()

    # даты не позже отчётной — чтобы будущие плановые записи не искажали факт
    horizon = [d for d in active_days if d <= report_date.isoformat()]

    ops: dict[str, dict] = {}
    for row in rows:
        key = f"{row.section}||{row.op}"
        op = ops.setdefault(key, {
            "id": slugify(row.op) + "-" + str(len(ops)),
            "section": row.section,
            "name": row.op,
            "plan": 0.0,
            "fact": 0.0,
            "daily": {},
            "items": [],
        })
        fact = round(sum(v for d, v in row.daily.items() if d <= report_date.isoformat()), 2)
        op["plan"] += row.plan
        op["fact"] += fact
        for day, val in row.daily.items():
            if day <= report_date.isoformat():
                op["daily"][day] = round(op["daily"].get(day, 0.0) + val, 2)

        pct = round(fact / row.plan * 100, 1) if row.plan else 0.0
        op["items"].append({
            "sub": row.sub,
            "culture": row.culture,
            "farm": row.farm,
            "plan": round(row.plan, 2),
            "fact": fact,
            "left": round(max(row.plan - fact, 0), 2),
            "pct": pct,
            "status": status(pct),
            "planStart": row.plan_start.isoformat() if row.plan_start else None,
            "planEnd": row.plan_end.isoformat() if row.plan_end else None,
            "factStart": row.fact_start.isoformat() if row.fact_start else (
                min((d for d in row.daily if d <= report_date.isoformat()), default=None)),
            "factEnd": row.fact_end.isoformat() if row.fact_end else None,
            "daily": {d: v for d, v in sorted(row.daily.items()) if d <= report_date.isoformat()},
            "row": row.excel_row,
        })

    op_list = []
    for op in ops.values():
        op["plan"] = round(op["plan"], 2)
        op["fact"] = round(op["fact"], 2)
        op["left"] = round(max(op["plan"] - op["fact"], 0), 2)
        op["pct"] = round(op["fact"] / op["plan"] * 100, 1) if op["plan"] else 0.0
        op["status"] = status(op["pct"])
        starts = [i["factStart"] for i in op["items"] if i["factStart"]]
        p_starts = [i["planStart"] for i in op["items"] if i["planStart"]]
        p_ends = [i["planEnd"] for i in op["items"] if i["planEnd"]]
        op["factStart"] = min(starts) if starts else None
        op["planStart"] = min(p_starts) if p_starts else None
        op["planEnd"] = max(p_ends) if p_ends else None
        op["daily"] = dict(sorted(op["daily"].items()))
        op_list.append(op)

    op_list.sort(key=lambda o: (o["section"], -o["plan"]))

    totals_by_day = {}
    for op in op_list:
        for day, val in op["daily"].items():
            totals_by_day[day] = round(totals_by_day.get(day, 0.0) + val, 2)

    plan_total = round(sum(o["plan"] for o in op_list), 2)
    fact_total = round(sum(o["fact"] for o in op_list), 2)

    last_day = horizon[-1] if horizon else None
    last7 = horizon[-7:]

    return {
        "meta": {
            "title": "План-факт полевых работ",
            "season": str(report_date.year),
            "reportDate": report_date.isoformat(),
            "generatedAt": dt.datetime.now().replace(microsecond=0).isoformat(),
            "source": "Общий свод 2026",
        },
        "totals": {
            "plan": plan_total,
            "fact": fact_total,
            "left": round(max(plan_total - fact_total, 0), 2),
            "pct": round(fact_total / plan_total * 100, 1) if plan_total else 0.0,
            "lastDay": last_day,
            "lastDayFact": totals_by_day.get(last_day, 0.0) if last_day else 0.0,
            "week": round(sum(totals_by_day.get(d, 0.0) for d in last7), 2),
            "tempo": round(sum(totals_by_day.get(d, 0.0) for d in last7) / max(len(last7), 1), 1),
        },
        "days": horizon,
        "byDay": totals_by_day,
        "sections": sorted({o["section"] for o in op_list if o["section"]}),
        "farms": sorted({i["farm"] for o in op_list for i in o["items"] if i["farm"] != "—"}),
        "cultures": sorted({i["culture"] for o in op_list for i in o["items"] if i["culture"] != "—"}),
        "ops": op_list,
    }
