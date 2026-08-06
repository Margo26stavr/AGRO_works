# -*- coding: utf-8 -*-
"""
Сборка отчёта из «Общего свода» в два целевых артефакта:

  docs/data.json  — данные для мобильного отчёта (GitHub Pages)
  docs/otchet.xlsx — Excel-отчёт для руководства (кнопка «Скачать Excel»)

Запуск:
  python scripts/build_report.py --src data/ПланФакт2026.xlsm --out docs
  python scripts/build_report.py --src ... --date 2026-05-24   # отчёт на конкретную дату
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from parse_svod import build_model, parse

FONT = "Arial"
INK = "16211C"
LEAF = "2E6B41"
STRAW = "B57E0C"
CLAY = "A63D28"
LINE = "D3D8CC"

RU_MON = ["янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]
RU_DOW = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]

thin = Side(style="thin", color=LINE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row: int, last_col: int, height: int = 30):
    for c in range(1, last_col + 1):
        cell = ws.cell(row, c)
        cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = height


def title(ws, text: str, last_col: int, subtitle: str = ""):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=INK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.row_dimensions[1].height = 22
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name=FONT, size=9, italic=True, color="6B7469")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)


def body(ws, r1: int, r2: int, last_col: int, size: int = 10):
    for r in range(r1, r2 + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT, size=size)
            cell.border = BOX


# ---------------------------------------------------------------- лист «Данные»
def sheet_data(wb: Workbook, model: dict) -> int:
    ws = wb.create_sheet("Данные")
    head = ["Раздел", "Вид работ", "Подоперация", "Культура", "Предприятие",
            "План, га", "Факт, га", "Остаток, га", "Выполнение, %",
            "Начало ПЛАН", "Окончание ПЛАН", "Начало ФАКТ", "Окончание ФАКТ", "Строка свода"]
    title(ws, "ДАННЫЕ — плоская выгрузка «Общего свода»", len(head),
          "Значения выгружены из свода. Столбцы H и I — формулы; остальное менять только в своде.")
    ws.append([])
    for i, h in enumerate(head, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(head))

    r = 5
    for op in model["ops"]:
        for it in op["items"]:
            ws.cell(r, 1, op["section"])
            ws.cell(r, 2, op["name"])
            ws.cell(r, 3, it["sub"])
            ws.cell(r, 4, it["culture"])
            ws.cell(r, 5, it["farm"])
            ws.cell(r, 6, it["plan"])
            ws.cell(r, 7, it["fact"])
            ws.cell(r, 8, f"=F{r}-G{r}")
            ws.cell(r, 9, f'=IFERROR(G{r}/F{r},"")')
            for col, key in ((10, "planStart"), (11, "planEnd"), (12, "factStart"), (13, "factEnd")):
                if it[key]:
                    ws.cell(r, col, dt.date.fromisoformat(it[key]))
            ws.cell(r, 14, it["row"])
            r += 1
    last = r - 1

    body(ws, 5, last, len(head), size=9)
    for rr in range(5, last + 1):
        for cc in (6, 7, 8):
            ws.cell(rr, cc).number_format = "#,##0"
        ws.cell(rr, 9).number_format = "0%"
        for cc in range(10, 14):
            ws.cell(rr, cc).number_format = "DD.MM.YYYY"
    widths = [22, 26, 24, 26, 16, 11, 11, 11, 12, 13, 14, 13, 14, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:N{last}"
    return last


# ------------------------------------------------------------ лист «Выработка»
def sheet_daily(wb: Workbook, model: dict) -> int:
    ws = wb.create_sheet("Выработка")
    head = ["Дата", "День", "Вид работ", "Культура", "Предприятие", "Га за день"]
    title(ws, "ВЫРАБОТКА ПО ДНЯМ — длинная таблица", len(head),
          "Одна строка = один день по одному срезу. Источник для сводных формул.")
    ws.append([])
    for i, h in enumerate(head, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(head))

    records = []
    for op in model["ops"]:
        for it in op["items"]:
            for day, val in it["daily"].items():
                records.append((day, op["name"], it["culture"], it["farm"], val))
    records.sort()

    r = 5
    for day, op_name, culture, farm, val in records:
        d = dt.date.fromisoformat(day)
        ws.cell(r, 1, d).number_format = "DD.MM.YYYY"
        ws.cell(r, 2, RU_DOW[d.weekday()])
        ws.cell(r, 3, op_name)
        ws.cell(r, 4, culture)
        ws.cell(r, 5, farm)
        ws.cell(r, 6, val).number_format = "#,##0"
        r += 1
    last = r - 1

    body(ws, 5, last, len(head), size=9)
    for i, w in enumerate([12, 7, 26, 26, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{last}"
    return last


# --------------------------------------------------------------- лист «Сводка»
def sheet_summary(wb: Workbook, model: dict, data_last: int):
    ws = wb.create_sheet("Сводка", 0)
    head = ["Вид работ", "План, га", "Факт, га", "Остаток, га", "Выполнение", "Статус", "Позиций"]
    rd = dt.date.fromisoformat(model["meta"]["reportDate"])
    title(ws, "ПЛАН-ФАКТ ПОЛЕВЫХ РАБОТ " + model["meta"]["season"], len(head),
          f"На {rd.strftime('%d.%m.%Y')} · собрано автоматически из листа «{model['meta']['source']}» · "
          f"все числа — формулы к листам «Данные» и «Выработка»")
    ws.append([])
    for i, h in enumerate(head, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(head))

    dref = f"Данные!$B$5:$B${data_last}"
    r = 5
    for op in model["ops"]:
        ws.cell(r, 1, op["name"])
        ws.cell(r, 2, f'=SUMIF({dref},$A{r},Данные!$F$5:$F${data_last})')
        ws.cell(r, 3, f'=SUMIF({dref},$A{r},Данные!$G$5:$G${data_last})')
        ws.cell(r, 4, f"=B{r}-C{r}")
        ws.cell(r, 5, f'=IFERROR(C{r}/B{r},0)')
        ws.cell(r, 6, f'=IF(E{r}>=0.995,"завершено",IF(E{r}>=0.7,"в графике",IF(E{r}>=0.35,"отставание","срыв срока")))')
        ws.cell(r, 7, f'=COUNTIF({dref},$A{r})')
        r += 1
    last = r - 1

    total = last + 1
    ws.cell(total, 1, "ИТОГО ПО ПРЕДПРИЯТИЮ")
    ws.cell(total, 2, f"=SUM(B5:B{last})")
    ws.cell(total, 3, f"=SUM(C5:C{last})")
    ws.cell(total, 4, f"=B{total}-C{total}")
    ws.cell(total, 5, f'=IFERROR(C{total}/B{total},0)')
    ws.cell(total, 7, f"=SUM(G5:G{last})")

    body(ws, 5, total, len(head))
    for rr in range(5, total + 1):
        for cc in (2, 3, 4):
            ws.cell(rr, cc).number_format = "#,##0"
        ws.cell(rr, 5).number_format = "0.0%"
        ws.cell(rr, 1).alignment = Alignment(vertical="center", wrap_text=True)
    for cc in range(1, len(head) + 1):
        c = ws.cell(total, cc)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INK)

    ws.conditional_formatting.add(
        f"E5:E{last}", DataBarRule(start_type="num", start_value=0, end_type="num",
                                   end_value=1, color=LEAF, showValue=True))
    for rule, text in ((CellIsRule(operator="lessThan", formula=["0.35"],
                                   font=Font(name=FONT, size=10, bold=True, color=CLAY)), "срыв"),
                       (CellIsRule(operator="between", formula=["0.35", "0.6999"],
                                   font=Font(name=FONT, size=10, color=STRAW)), "отставание")):
        ws.conditional_formatting.add(f"F5:F{last}", rule)

    for i, w in enumerate([34, 12, 12, 13, 13, 16, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


# -------------------------------------------------------------- лист «По дням»
def sheet_bydays(wb: Workbook, model: dict, daily_last: int, days_back: int = 30):
    ws = wb.create_sheet("По дням")
    days = [dt.date.fromisoformat(d) for d in model["days"][-days_back:]]
    ncols = 2 + len(days) + 1
    rd = dt.date.fromisoformat(model["meta"]["reportDate"])
    title(ws, f"ВЫРАБОТКА ПО ДНЯМ, га — последние {len(days)} рабочих дней", ncols,
          f"На {rd.strftime('%d.%m.%Y')} · «Накоплено» — нарастающим итогом с начала сезона")

    ws.cell(4, 1, "Вид работ")
    ws.cell(4, 2, "Накоплено")
    for i, d in enumerate(days):
        c = ws.cell(4, 3 + i, f"{RU_DOW[d.weekday()]}\n{d.strftime('%d.%m')}")
    ws.cell(4, 3 + len(days), "Итого\nза период")
    style_header(ws, 4, ncols, height=34)

    # служебная строка с датами — на неё ссылаются SUMIFS
    for i, d in enumerate(days):
        cell = ws.cell(3, 3 + i, d)
        cell.number_format = "DD.MM.YYYY"
        cell.font = Font(name=FONT, size=1, color="FFFFFF")
    ws.row_dimensions[3].hidden = True

    dates_rng = f"Выработка!$A$5:$A${daily_last}"
    ops_rng = f"Выработка!$C$5:$C${daily_last}"
    vals_rng = f"Выработка!$F$5:$F${daily_last}"

    r = 5
    for op in model["ops"]:
        ws.cell(r, 1, op["name"])
        ws.cell(r, 2, f'=SUMIF({ops_rng},$A{r},{vals_rng})')
        for i in range(len(days)):
            col = get_column_letter(3 + i)
            ws.cell(r, 3 + i,
                    f'=IF(SUMIFS({vals_rng},{ops_rng},$A{r},{dates_rng},{col}$3)=0,"",'
                    f'SUMIFS({vals_rng},{ops_rng},$A{r},{dates_rng},{col}$3))')
        c0, c1 = get_column_letter(3), get_column_letter(2 + len(days))
        ws.cell(r, 3 + len(days), f"=SUM({c0}{r}:{c1}{r})")
        r += 1
    last = r - 1

    total = last + 1
    ws.cell(total, 1, "ИТОГО ЗА ДЕНЬ")
    for cc in range(2, ncols + 1):
        L = get_column_letter(cc)
        ws.cell(total, cc, f"=SUM({L}5:{L}{last})")

    body(ws, 5, total, ncols, size=9)
    for rr in range(5, total + 1):
        for cc in range(2, ncols + 1):
            ws.cell(rr, cc).number_format = "#,##0"
            ws.cell(rr, cc).alignment = Alignment(horizontal="center")
        ws.cell(rr, 1).font = Font(name=FONT, size=9)
    for cc in range(1, ncols + 1):
        c = ws.cell(total, cc)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INK)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    for i in range(len(days)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 7.5
    ws.column_dimensions[get_column_letter(3 + len(days))].width = 11
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False


# ----------------------------------------------------------------- лист «Сроки»
def sheet_terms(wb: Workbook, model: dict, data_last: int):
    ws = wb.create_sheet("Сроки")
    head = ["Вид работ", "Культура", "Предприятие",
            "Начало ПЛАН", "Окончание ПЛАН", "Начало ФАКТ", "Окончание ФАКТ",
            "Сдвиг старта, дн.", "Выполнение", "Комментарий"]
    rd = dt.date.fromisoformat(model["meta"]["reportDate"])
    title(ws, "СРОКИ ВЫПОЛНЕНИЯ РАБОТ", len(head),
          f"На {rd.strftime('%d.%m.%Y')} · положительный сдвиг = вышли в поле позже агротехнического срока")
    ws.append([])
    for i, h in enumerate(head, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(head))

    r = 5
    src = 5
    for op in model["ops"]:
        for it in op["items"]:
            ws.cell(r, 1, op["name"])
            ws.cell(r, 2, it["culture"])
            ws.cell(r, 3, it["farm"])
            for col, key in ((4, "planStart"), (5, "planEnd"), (6, "factStart"), (7, "factEnd")):
                if it[key]:
                    ws.cell(r, col, dt.date.fromisoformat(it[key])).number_format = "DD.MM.YYYY"
            ws.cell(r, 8, f'=IF(AND(D{r}<>"",F{r}<>""),F{r}-D{r},"")')
            ws.cell(r, 9, f"=Данные!I{src}")
            ws.cell(r, 10,
                    f'=IF(I{r}>=0.995,"закрыто",'
                    f'IF(AND(E{r}<>"",E{r}<TODAY()),"агросрок истёк, не закрыто",'
                    f'IF(F{r}="","не начинали","в работе")))')
            r += 1
            src += 1
    last = r - 1

    body(ws, 5, last, len(head), size=9)
    for rr in range(5, last + 1):
        ws.cell(rr, 8).number_format = "0"
        ws.cell(rr, 8).alignment = Alignment(horizontal="center")
        ws.cell(rr, 9).number_format = "0%"
    ws.conditional_formatting.add(
        f"H5:H{last}", CellIsRule(operator="greaterThan", formula=["0"],
                                  font=Font(name=FONT, size=9, bold=True, color=CLAY)))
    ws.conditional_formatting.add(
        f"I5:I{last}", DataBarRule(start_type="num", start_value=0, end_type="num",
                                   end_value=1, color=LEAF, showValue=True))

    for i, w in enumerate([28, 26, 15, 13, 14, 13, 14, 15, 12, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:J{last}"
    ws.sheet_view.showGridLines = False


def build_xlsx(model: dict, path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    data_last = sheet_data(wb, model)
    daily_last = sheet_daily(wb, model)
    sheet_summary(wb, model, data_last)
    sheet_bydays(wb, model, daily_last)
    sheet_terms(wb, model, data_last)
    wb.move_sheet("По дням", offset=-2)
    wb.move_sheet("Сроки", offset=-2)
    wb.save(path)


def build_offline(out: Path, payload: str):
    """Один HTML-файл с вшитыми данными — можно переслать в мессенджере, интернет не нужен."""
    page = out / "index.html"
    if not page.exists():
        return
    html = page.read_text(encoding="utf-8")
    safe = payload.replace("</", "<\\/")
    # тег вставляется сразу после <body> — до основного скрипта, иначе он его не найдёт
    tag = f'<body>\n<script id="embedded" type="application/json">{safe}</script>'
    (out / "otchet-offline.html").write_text(html.replace("<body>", tag, 1), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser(description="Сборка мобильного отчёта и Excel из «Общего свода»")
    ap.add_argument("--src", default="data/ПланФакт2026.xlsm")
    ap.add_argument("--out", default="docs")
    ap.add_argument("--date", help="отчётная дата ГГГГ-ММ-ДД (по умолчанию — последний день с выработкой)")
    args = ap.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    report_date = dt.date.fromisoformat(args.date) if args.date else None
    model = build_model(parse(args.src), report_date)

    payload = json.dumps(model, ensure_ascii=False, separators=(",", ":"))
    (out / "data.json").write_text(payload, encoding="utf-8")
    build_xlsx(model, out / "otchet.xlsx")
    build_offline(out, payload)

    t = model["totals"]
    print(f"Отчётная дата : {model['meta']['reportDate']}")
    print(f"Видов работ   : {len(model['ops'])}")
    print(f"План / факт   : {t['plan']:,.0f} / {t['fact']:,.0f} га  ({t['pct']}%)")
    print(f"Записано      : {out/'data.json'}, {out/'otchet.xlsx'}, {out/'otchet-offline.html'}")


if __name__ == "__main__":
    main()
